import os, time, flask, openpyxl, string
from os.path import join, dirname, realpath
from openpyxl import load_workbook
from flask import Flask
from flask import Blueprint, render_template, request, redirect, url_for, flash, sessions, session, send_from_directory
from flask_mysqldb import MySQL
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import *
from flask import app

def create_app():
    # create and configure the app
    #app = Flask(__name__, instance_relative_config=True)
    app = Flask(__name__)
    app.secret_key = b'5t759f9$gfdhf0478y87^4#5gq8*3nft8503#mgtrhsuooer9'

    #app.config['MYSQL_HOST'] = 'localhost'
    #app.config['MYSQL_USER'] = 'root'
    #app.config['MYSQL_PASSWORD'] = ''
    #app.config['MYSQL_DB'] = 'Splitter'

    app.config['MYSQL_HOST'] = 'Timdeplatvis111.mysql.pythonanywhere-services.com'
    app.config['MYSQL_USER'] = 'Timdeplatvis111'
    app.config['MYSQL_PASSWORD'] = 'CdYudQM75q7DxHh'
    app.config['MYSQL_DB'] = 'Timdeplatvis111$Splitter'

    mysql = MySQL(app)

    #16 Megabytes is se maximum file size
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

    app.config.from_mapping(
        SECRET_KEY = b'5t759f9$gfdhf0478y87^4#5gq8*3nft8503#mgtrhsuooer9',
        UPLOADED_FILES = 'files/',
        UPLOADS_PATH = join(dirname(realpath(__file__)), 'files/'),
    )

    ALLOWED_EXTENSIONS = set(['xlsx'])
    def allowed_file(filename):
        return '.' in filename and \
        filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

    @app.route('/')
    def index():
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM post")
        posts = cur.fetchall()
        return render_template('index.html', posts=posts)

    @app.route('/feedback', methods=['POST', 'GET'])
    def feedback():
        try: 
            username = request.form['username']
            content = request.form['content']

            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO post(username, content) VALUES (%s, %s)", (username, content))
            mysql.connection.commit()
            cur.close()
            flash ("Done")
            return redirect("/")
            return render_template('index.html')
        except:
            return redirect("/")
            return render_template('index.html')

    @app.route('/upload', methods=['POST', 'GET'])
    def upload():
        try: 
            request.files['uploadedfiles']
            yeet = request.files.getlist('uploadedfiles')
            filenamen = []
            test = 'yeeticus'
            for file in yeet:     
                files = request.files.to_dict()
                filename = secure_filename(file.filename)
                filenamen.append(filename)

                if file and allowed_file(filename):
                    file.save(os.path.join(app.config['UPLOADS_PATH'], filename)) 
                else:
                    flash(filenamen, 'fileerror')
                    return render_template('index.html')

            session['filenamen[]'] = filenamen
            return render_template('convert.html', filenamen=session['filenamen[]'])    
        except KeyError:
            flash ('error')
            return redirect("/")
            return render_template('index.html')

    @app.route('/convert', methods=['GET', 'POST'])
    def convert():
        try:
            filenamen = session.get('filenamen[]')

            filenaam1 = filenamen[0]
            filenaam2 = filenamen[1]

            workbook1 = load_workbook(filename=f'/files/{filenaam1}')
            workbook2 = load_workbook(filename=f'/files/{filenaam2}')

            column1 = request.form['column1']
            column2 = request.form['column2']

            option = request.form['option']

            column1copy = request.form['column1copy']
            column2copy = request.form['column2copy']

            sheet1 = workbook1.active
            sheet2 = workbook2.active

            column1copy = int(column1copy)
            column2copy = int(column2copy)

            column1 = int(column1)
            column2 = int(column2)

            values = dict()
            for index, letter in enumerate(string.ascii_lowercase):
                values[index] = letter
                print(values[index])

            sheet1column = values[column1]
            sheet2column = values[column2]

            #column1 = de eerste column die je wilt matchen uit de form (filenaam1)
            #column2 = de tweede column die je wilt matchen uit de form (filenaam2)

            #option = de file waarvan je de values wilt kopieren uit de form

            #column1copy = de column waarvan je de values wilt kopieren uit de form
            #column2copy = de column waar naar je de values wilt kopieren uit de form

            t = time.process_time()
            loops = 0
            if option == 'file0':
                for cell in sheet1[sheet1column]:
                    I = cell.row
                    sheet1value = sheet1.cell(row=I, column=column1).value
                    #print(sheet1Value)
                    for cell in sheet2[sheet2column]:
                        E = cell.row
                        sheet2value = sheet2.cell(row=E, column=column2).value
                        if sheet1value == sheet2value:
                            sheet1copyvalue = sheet1.cell(row=I, column=column1copy).value
                            sheet2.cell(row=E, column=column2copy, value=sheet1copyvalue)
                            print(sheet1copyvalue)
                        else:
                            loops += 1
            else:
                for cell in sheet2[sheet2column]:
                    I = cell.row
                    sheet2value = sheet2.cell(row=I, column=column2).value
                    #print(sheet1Value)
                    for cell in sheet1[sheet1column]:
                        E = cell.row
                        sheet1value = sheet1.cell(row=E, column=column1).value
                        if sheet2value == sheet1value:
                            sheet2copyvalue = sheet2.cell(row=I, column=column1copy).value
                            sheet1.cell(row=E, column=column2copy, value=sheet2copyvalue)
                            print(column2copy)
                            print(sheet2copyvalue)
                        else:
                            loops += 1

            #VERGEET DIT NIET TE VERANDEREN
            workbook1.save(f'/files/{filenaam1}')
            workbook2.save(f'/files/{filenaam2}')
            elapsed_time = time.process_time() - t

            flash(elapsed_time, 'time')
            flash(loops, 'loops')

            return redirect(url_for('uploaded_file', filename=filenaam1))
            #return render_template('index.html', filenamen=session['filenamen[]'])

        except:
            return redirect("/")
            return render_template('index.html')
            
    #TODO: dit
    @app.route('/uploads/<filename>')
    def uploaded_file(filename):
        #return send_from_directory(app.config['UPLOADS_PATH'], filetouser)
        return redirect("/")
        return render_template('index.html')

    #Dit zorgt er dus voor dat de app uiteindelijk runt, alle gegevens zitten in app
    return app


