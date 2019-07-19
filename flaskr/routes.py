import os, time, flask, string, MySQLdb, openpyxl, wtforms, jinja2

from flask import Flask, Blueprint, render_template, request, redirect, url_for, flash, sessions, session, send_from_directory, send_file

from flaskr import app, allowed_file, flask_bcrypt, db, bcrypt, models

from os.path import join, dirname, realpath

from flask_sqlalchemy import SQLAlchemy

from openpyxl import load_workbook

from werkzeug.utils import *
from werkzeug.wrappers import BaseRequest
from werkzeug.wsgi import responder
from werkzeug.exceptions import HTTPException, NotFound

from forms import RegistrationForm, LoginForm, PostForm
from flaskr.models import User, Post
from flask_bcrypt import Bcrypt
from flask_login import login_user, current_user, logout_user, login_required
from flask_sqlalchemy import *
from sqlalchemy import *

from github import Github

from string import ascii_lowercase
import itertools

#Github configuration, for automatic update checking.
f = open("githublogin.txt", "r")
for x in f:
    username = f.readline()
    password = f.readline()

g = Github(username, password)

#For the exception catchers, for some reason they need a value first.
a, b, c, d = '', '', '', ''

@app.route('/', methods=['GET', 'POST'])
@app.route('/page/<int:page>', methods=['GET', 'POST'])
def index(page=1):
    try:
        registerform = RegistrationForm()
        if registerform.validate_on_submit():
            checkUsername = registerform.username.data
            checkEmail = registerform.email.data
            hashed_password = bcrypt.generate_password_hash(registerform.password.data).decode('utf-8')
            user = User(username=registerform.username.data, email=registerform.email.data, password=hashed_password)
            usernameExists = db.session.query(db.session.query(User).filter_by(username=checkUsername).exists()).scalar()
            emailExists = db.session.query(db.session.query(User).filter_by(email=checkEmail).exists()).scalar()
            if usernameExists or emailExists:
                message = 'That username or email is already taken'
                flash(str(message), 'loginError')
                return redirect("/")
                return render_template('index.html', loginError=loginError)
            else:
                db.session.add(user)
                db.session.commit()
                message = 'Registration succesfull!'
                flash(str(message), 'loginError')
                return redirect("/")
                return render_template('index.html', loginError=loginError)

            return redirect("/")
            return render_template('index.html', loginError=loginError)

        loginform = LoginForm()
        if loginform.validate_on_submit():
            user = User.query.filter_by(email=loginform.email.data).first()
            if user and bcrypt.check_password_hash(user.password, loginform.password.data):
                login_user(user, remember=loginform.remember.data)
                #next_page = request.args.get('next')
                #return redirect(next_page) if next_page else redirect(url_for('index'))
                return redirect(url_for('/'))
            else:
                message = 'Invalid login, please check your login values and try again'
                flash(str(message), 'loginError')
                return redirect("/")
                return render_template('index.html', loginError=loginError)

        if current_user.is_authenticated:
            userfolder = current_user.username
            converteduserfiles = []
            userfiles = []

            path = f'files/{userfolder}/'

            if (os.path.exists(f'files/{userfolder}/converted')):
                pathtoconverted = f'files/{userfolder}/converted'
            else:
                if not (os.path.exists(f'files/{userfolder}')):
                    os.mkdir(f'files/{userfolder}')
                os.mkdir(f'files/{userfolder}/converted')
                pathtoconverted = f'files/{userfolder}/converted'

            for filename in os.listdir(path):
                if os.path.isfile and filename != 'converted':
                    userfiles.append(filename)

            for filename in os.listdir(pathtoconverted):
                if os.path.isfile:
                    converteduserfiles.append(filename)
        else:
            filename = ''
            path = ''
            userfiles = '', ''
            converteduserfiles = ''
            pathtoconverted = ''
            session['filename'] = filename
            session['path'] = path
            session['userfiles[]'] = userfiles
            session['converteduserfiles[]'] = converteduserfiles
            session['pathtoconverted'] = pathtoconverted

        session['filename'] = filename
        session['path'] = path
        session['userfiles[]'] = userfiles
        session['converteduserfiles[]'] = converteduserfiles
        session['pathtoconverted'] = pathtoconverted

        postform = PostForm()
        if postform.validate_on_submit():
            post = Post(title=postform.title.data, content=postform.content.data, author=current_user)
            db.session.add(post)
            db.session.commit()
            flash('Your post has been created!', 'success')
            return redirect(url_for('index'))

        RESULTS_PER_PAGE = 5
        #posts = Post.query.all()
        #models.Post.query.paginate(page, per_page, error_out=False)
        #posts = Post.query.order_by(Post.id.title()).paginate(page,per_page,error_out=False)
        posts = models.Post.query.paginate(page, RESULTS_PER_PAGE, False)
        num = int(ceil(float(posts.total) / RESULTS_PER_PAGE)) + 1

        environment = jinja2.Environment(os)
        environment.filters['os'] = os

        #{% for post in posts|sort(attribute='date_posted', reverse=true) %}

        #VERGEET DIT NIET TE VERANDEREN IN BETWEEN RELEASES TIM
        currentVersion = "Excelsplitter version 1.0.2"

        repo = g.get_repo("timdeplatvis111/ExcelSplitter")
        print(repo.name)
        #repo.compare
        latestRelease = repo.get_latest_release()
        if latestRelease.title == currentVersion:
            uptodate = 'true'
        else:
            uptodate = 'false'

        return render_template('index.html', title='Account', loginform=loginform, registerform=registerform, postform=postform, posts=posts, number_of_pages=num, userfiles=session['userfiles[]'], path=session['path'], filename=session['filename'], pathtoconverted=session['pathtoconverted'], converteduserfiles=session['converteduserfiles[]'], os=os, uptodate=uptodate)

    #All exception catchers, most of these will never happen but they're there just to be sure.
    except KeyError as a:
        flash(str(a), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        #message = 'You broke my webapp somehow, if this is a recurring error then please contact the developer'
        #flash(str(message), 'error')
        return redirect("/")
        #return render_template('index.html', error=error)
        return render_template('index.html')
        session.pop('_flashes', None)
"""
@app.route(, methods=['GET', 'POST'])
def userfiles():
    filename = session.get('filename')
    path = session.get('path')
    userfiles = session.get('userfiles[]')

    for index, filename in enumerate(userfiles):
        print('nice')
    return send_from_directory(f'../{path}', userfiles[index], as_attachment=True)
"""

@app.route('/files/<filename>', methods=['GET', 'POST'])
def files(filename):
    try: 
        path = session.get('path')
        userfiles = session.get('userfiles[]')

        return send_from_directory(f'../{path}', filename, as_attachment=True)

    except KeyError as d:
        flash(str(d), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

@app.route('/files/converted/<filename>', methods=['GET', 'POST'])
def files2(filename):
    try: 
        pathtoconverted = session.get('pathtoconverted')
        converteduserfiles = session.get('converteduserfiles[]')

        return send_from_directory(f'../{pathtoconverted}', filename, as_attachment=True)

    except KeyError as d:
        flash(str(d), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/upload', methods=['POST', 'GET'])
def upload():
    try: 
        #Takes the files from the forms
        request.files['uploadedfile1']
        request.files['uploadedfile2']

        userfolder = current_user.username

        #Creates a list from the files so we can loop through them later
        yeet1 = request.files.getlist('uploadedfile1')
        yeet2 = request.files.getlist('uploadedfile2')

        yeet3 = yeet1 + yeet2
        filenamen = []

        #Loops through the filelist
        for file in yeet3:     
            files = request.files.to_dict()
            filename = secure_filename(file.filename)
            filenamen.append(filename)

            #Checks if the uploaded file is an .xlxs file
            if file and allowed_file(filename):
                if (os.path.exists(f'files/{userfolder}')):
                    file.save(f'files/{userfolder}/{filename}')
                else:
                    os.makedirs(f'files/{userfolder}')
                    file.save(f'files/{userfolder}/{filename}')
            else:
                fileerror= 'fuck'
                flash('yeet', 'fileerror')
                return render_template('index.html', fileerror=fileerror)

        if len(filenamen) is not 2:
            flash('Please upload 2 Excel files', 'error')
            return render_template('convert.html', error=error)    
        else:
            #Puts the filenames into a session for use in the next route
            session['filenamen[]'] = filenamen
            return render_template('convert.html', filenamen=session['filenamen[]'])    

    except KeyError as d:
        flash(str(d), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

@app.route('/brupload', methods=['POST', 'GET'])
def brupload():
    try: 
        request.files['bruploadedfile']
        userfolder = current_user.username

        yeet4 = request.files.getlist('bruploadedfile')

        brfilenamen = []

        for file in yeet4:
            files = request.files.to_dict()
            filename = secure_filename(file.filename)
            brfilenamen.append(filename)

            print(brfilenamen[0])

            #Checks if the uploaded file is an .xlxs file
            if file and allowed_file(filename):
                if (os.path.exists(f'files/{userfolder}/br')):
                    file.save(f'files/{userfolder}/br/{filename}')
                else:
                    os.makedirs(f'files/{userfolder}/br')
                    file.save(f'files/{userfolder}/br/{filename}')
            else:
                fileerror= 'fuck'
                flash('yeet', 'fileerror')
                return render_template('index.html', fileerror=fileerror)

        session['brfilenamen[]'] = brfilenamen
        return render_template('brconvert.html', brfilenamen=session['brfilenamen[]'])  

    except KeyError as d:
        flash(str(d), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

@app.route('/brconvert', methods=['GET', 'POST'])
def brconvert():
    try:
        def iter_all_strings():
            for size in itertools.count(1):
                for letter in itertools.product(ascii_lowercase, repeat=size):
                    yield "".join(letter)

        values = dict()
        randomindex = 0

        for letter in itertools.islice(iter_all_strings(), 30):
            randomindex +=1
            values[randomindex] = letter

        userfolder = current_user.username
        
        brfilenamen = session.get('brfilenamen[]')

        filenaam1 = brfilenamen[0]

        print(brfilenamen[0])

        workbook1 = load_workbook(filename=(f"files/{userfolder}/br/{filenaam1}"))
        sheet1 = workbook1.active

        column1 = request.form['column1']

        column1 = int(column1)

        sheet1column = values[column1]

        br = '<br> '
        closebr = ' <br />'

        for cell in sheet1[sheet1column]:
            I = cell.row
            brplacedata = sheet1.cell(row=I, column=column1).value
            brplacedata = br + brplacedata + closebr
            sheet1.cell(row=I, column=column1, value=brplacedata)
        
        workbook1.save(f'files/{userfolder}/{filenaam1}')

        return send_from_directory(f'../files/{userfolder}', filenaam1, as_attachment=True)

    except KeyError as d:
        flash(str(d), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except ValueError as c:
        flash(str(c), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)
        session.pop('_flashes', None)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

@app.route('/convert', methods=['GET', 'POST'])
def convert():
    try:
        def iter_all_strings():
            for size in itertools.count(1):
                for letter in itertools.product(ascii_lowercase, repeat=size):
                    yield "".join(letter)

        values = dict()
        randomindex = 0
        for letter in itertools.islice(iter_all_strings(), 30):
            randomindex +=1
            values[randomindex] = letter

        #Sets the userfolder equal to the logged in user's username.
        userfolder = current_user.username

        #Gets the filenamen list from the session
        filenamen = session.get('filenamen[]')

        #Splits the array into 2 seperate variables
        filenaam1 = filenamen[0]
        filenaam2 = filenamen[1]

        #column1 = The first column you want to compare, generated from the form (filenaam1)
        #column2 = The second column you want to compare, generated from the form (filenaam2)
        column1 = request.form['column1']
        column2 = request.form['column2']

        #option = Variable for deciding which file you want to copy values from
        option = request.form['option']

        #column1copy = The column where it copies the values from 
        #column2copy = The column where it copies the values to
        column1copy = request.form['column1copy']
        column2copy = request.form['column2copy']

        #part1column = The first position for the string matching system
        #part2column = The second position for the string matching system

        #Example: part1column = 5, part2column = 7                        
        #With these values it matches the 6th character (counting from 1) to the 8th character (counting from 1)
        partcolumn1 = request.form['partcolumn1']
        partcolumn2 = request.form['partcolumn2']

        #Sets both workbooks as active, so we can manipulate them using openpyxl
        workbook1 = load_workbook(filename=(f"files/{userfolder}/{filenaam1}"))
        sheet1 = workbook1.active

        workbook2 = load_workbook(filename=(f"files/{userfolder}/{filenaam2}"))
        sheet2 = workbook2.active

        #Converts these values to integers for later use
        column1copy = int(column1copy)
        column2copy = int(column2copy)

        column1 = int(column1)
        column2 = int(column2)

        #Checks if these values are empty, then sets them to a high value to prevent errors
        if partcolumn1 == '' or partcolumn2 == '':
            partcolumn1 = 0 
            partcolumn2 = 100000
            partcolumn1 = int(partcolumn1)
            partcolumn2 = int(partcolumn2)
        else:
            partcolumn1 = int(partcolumn1)
            partcolumn2 = int(partcolumn2)

        sheet1column = values[column1]
        sheet2column = values[column2]

        keepdataoption = request.form['keepdataoption']

        colourcells = request.form['colourcells']
        color1 = request.form['color1']

        #Sets the colour of the cell coloring feature.
        myColour = openpyxl.styles.colors.Color(rgb=color1)
        colourFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=myColour)

        #Defines some variables for measuring the time, amount of loops, and the amount of skipped loops
        t = time.process_time()
        loops = 0
        errorloops = 0

        #This if statement can definitely be removed somehow, writing this code twice is really inefficient. 
        if option == 'file0':
            filename =  filenaam1

            for cell in sheet1[sheet1column]:
                try:
                    filename = filenaam2
                    I = cell.row
                    sheet1value = sheet1.cell(row=I, column=column1).value
                    print(sheet1value)
                    #time.sleep(5)
                    if sheet1value != None:
                        try:
                            sheet1value = sheet1value[partcolumn1:partcolumn2]
                        except:
                            print("Out of range")
                            pass

                    for cell in sheet2[sheet2column]:
                        E = cell.row
                        sheet2value = sheet2.cell(row=E, column=column2).value
                        print(sheet2value)
                        #time.sleep(5)
                        if sheet2value != None:
                            try:
                                sheet2value = sheet2value[partcolumn1:partcolumn2]
                            except: 
                                print("Out of range")
                                pass

                        if sheet1value == sheet2value:
                            sheet1copyvalue = sheet1.cell(row=I, column=column1copy).value
                            if colourcells == 'colourcells2':
                                sheet2.cell(row=E, column=column2copy).fill = colourFill
                            if keepdataoption == 'keepdata1':
                                keepdatacell = sheet2.cell(row=E, column=column2copy).value
                                try:
                                    keepdatacell = keepdatacell + sheet1copyvalue
                                    sheet2.cell(row=E, column=column2copy, value=keepdatacell)
                                    loops +=1
                                except: 
                                    sheet2.cell(row=E, column=column2copy, value=sheet1copyvalue)
                                    loops +=1
                            else:
                                sheet2.cell(row=E, column=column2copy, value=sheet1copyvalue)
                                loops +=1
                        else:
                            loops +=1
                except:
                    print('F')
                    print('sheet1value')
                    print(sheet1value)
                    print('sheet2value')
                    print(sheet2value)
                    errorloops +=1

        elif option == 'file1':
            filename = filenaam2

            for cell in sheet2[sheet2column]:
                try:
                    I = cell.row
                    sheet2value = sheet2.cell(row=I, column=column2).value
                    print(sheet2value)
                    #time.sleep(5)
                    if sheet2value != None:
                        try:
                            sheet2value = sheet2value[partcolumn1:partcolumn2]
                        except:
                            print("Out of range")
                            pass

                    for cell in sheet1[sheet1column]:
                        E = cell.row
                        sheet1value = sheet1.cell(row=E, column=column1).value
                        print(sheet1value)
                        #time.sleep(5)
                        if sheet1value != None:
                            try:
                                sheet1value = sheet1value[partcolumn1:partcolumn2]
                            except:
                                print("Out of range")
                                pass

                        if sheet2value == sheet1value:
                            sheet1copyvalue = sheet2.cell(row=I, column=column1copy).value
                            if colourcells == 'colourcells2':
                                sheet1.cell(row=E, column=column2copy).fill = colourFill
                            if keepdataoption == 'keepdata1':
                                keepdatacell = sheet1.cell(row=E, column=column2copy).value
                                try:
                                    keepdatacell = keepdatacell + sheet1copyvalue
                                    sheet1.cell(row=E, column=column2copy, value=keepdatacell)
                                    loops +=1
                                except:
                                    sheet1.cell(row=E, column=column2copy, value=sheet1copyvalue)
                                    loops +=1
                            else:
                                sheet1.cell(row=E, column=column2copy, value=sheet1copyvalue)
                                loops +=1
                        else:
                            loops +=1
                except:
                    print('F')
                    print('sheet1value')
                    print(sheet1value)
                    print('sheet2value')
                    print(sheet2value)
                    errorloops +=1

        #Saves the files to our userfolder
        if option == 'file0':
            if (os.path.exists(f'files/{userfolder}')):
                workbook1.save(f'files/{userfolder}/{filenaam1}')
                workbook2.save(f'files/{userfolder}/{filenaam2}')
            else:
                os.makedirs(f'files/{userfolder}')
                workbook1.save(f'files/{userfolder}/{filenaam1}')
                workbook2.save(f'files/{userfolder}/{filenaam2}')

        elif option == 'file1':
            if (os.path.exists(f'files/{userfolder}')):
                workbook1.save(f'files/{userfolder}/{filenaam1}')
                workbook2.save(f'files/{userfolder}/{filenaam2}')
            else:
                os.makedirs(f'files/{userfolder}')
                workbook1.save(f'files/{userfolder}/{filenaam1}')
                workbook2.save(f'files/{userfolder}/{filenaam2}')

        #Measured the time it takes for the program to go through all the code above
        elapsed_time = time.process_time() - t

        #Renders the amount of loops, errorloops and the amount of time it took to go through all the code above
        flash(elapsed_time, 'time')
        flash(loops, 'loops')
        flash(errorloops, 'errorloops')

        if (os.path.exists(f'files/{userfolder}/converted')):
            print('Converted yeet')
        else:
            os.mkdir(f'files/{userfolder}/converted')

        if option == 'file0':
            workbook2.save(f'files/{userfolder}/converted/{filenaam2}')
            return send_from_directory(f'../files/{userfolder}', filenaam2, as_attachment=True)
        elif option == 'file1':
            workbook1.save(f'files/{userfolder}/converted/{filenaam1}')
            return send_from_directory(f'../files/{userfolder}', filenaam1, as_attachment=True)

    #Exception catchers, just to be sure
    except KeyError as a:
        flash(str(a), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except NameError as b:
        flash(str(b), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except ValueError as c:
        flash(str(c), 'valueError')
        return redirect("/")
        return render_template('index.html', valueError=valueError)

    except TypeError as f:
        flash(str(f), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

    except:
        message = 'An error was detected, please try again'
        flash(str(message), 'error')
        return redirect("/")
        return render_template('index.html', error=error)

        session.pop('_flashes', None) 

