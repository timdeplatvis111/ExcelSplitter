import openpyxl
from openpyxl import load_workbook

import time

AW18 = load_workbook(filename = 'C:\\Users\\timde\\Desktop\\KingsofIndigo\\InformatieKoppelen\\AW18.xlsx')
SS19 = load_workbook(filename = 'C:\\Users\\timde\\Desktop\\KingsofIndigo\\InformatieKoppelen\\SS19.xlsx')
Samen = load_workbook(filename = 'C:\\Users\\timde\\Desktop\\KingsofIndigo\\InformatieKoppelen\\Samenvoegen.xlsx')

AW18sheet = AW18.get_sheet_names
SS19sheet = SS19.get_sheet_names
Samensheet = Samen.get_sheet_names

AW18sheet = AW18.active
SS19sheet = SS19.active
Samensheet = Samen.active

max_row=Samensheet.max_row
#print(max_row)
Loops = 0

for cell in AW18sheet['A']:
  I = cell.row 
  KnummerAW18 = AW18sheet.cell(row=I, column=1).value
  print(KnummerAW18)
  for cell in Samensheet['B']:
    E = cell.row
    KnummerSamen = Samensheet.cell(row=E, column=2).value
    if KnummerAW18 == KnummerSamen:
      ModelInformation = AW18sheet.cell(row=I, column=5).value
      ModelSize = AW18sheet.cell(row=I, column=6).value 

      Samensheet.cell(row=E, column=56, value=ModelInformation)
      Samensheet.cell(row=E, column=57, value=ModelSize)
      print(ModelInformation)
      print(ModelSize)
      print(KnummerSamen)  
    else:
      Loops += 1

print (Loops)
Samen.save('Samenvoegen.xlsx')